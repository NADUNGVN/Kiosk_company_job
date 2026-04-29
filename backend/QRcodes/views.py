from django.shortcuts import render, redirect
from datetime import datetime
from django.urls import reverse
from backend.customer.models import  Employee
from django.http import HttpResponse
from config.settings import MEDIA_ROOT
from django.views.decorators.csrf import csrf_exempt
from mailmerge import MailMerge
from docx import Document
from django.conf import settings
import os
from openpyxl import load_workbook
from num2words import num2words
from django.contrib.auth import authenticate, login
import os
import subprocess
import win32com.client as win32 
import pythoncom  # Thêm import pythoncom
from openpyxl import Workbook
from backend.customer.models import KhachHang
def export(request,pk):
    if request.method == 'POST' and 'hidden_input' in request.POST:
        selected_option = int(request.POST.get('hidden_input'))
        so_thu_tu = request.POST.get('hidden_input_1')
        thong_tin_khach_hang = KhachHang.objects.get(ticket_number=so_thu_tu)
        thanh_pho = thong_tin_khach_hang.dia_chi.replace('\x00','').split(',')
        thanh_pho_chuan=thanh_pho[-1]
        print(thanh_pho_chuan)
        a=request.session['scan_data_1'] = {
                'ho_va_ten': thong_tin_khach_hang.name,
                'dia_chi':thong_tin_khach_hang.dia_chi,
                'ngay_sinh':thong_tin_khach_hang.ngay_sinh,
                'ticket_number': thong_tin_khach_hang.ticket_number,
                'so_cccd':thong_tin_khach_hang.so_cccd,
                'gioi_tinh':thong_tin_khach_hang.gioi_tinh,
                'ngay_cap':thong_tin_khach_hang.ngay_cap,
                'thanh_pho':thanh_pho_chuan,            
                # Add other fields as necessary
            }
        
        if selected_option == 2:
            return redirect('qrcode:hon-nhan-1')
        elif selected_option == 4:
            return redirect('qrcode:ho-tich-1')
        elif selected_option == 1:
            return redirect('qrcode:ket-hon-1')
        elif selected_option == 6:
            return redirect('qrcode:khai-sinh-1')
        elif selected_option == 7:
            return redirect('qrcode:khai-tu-1')
        else:
            return redirect('qrcode:upload')
            

    return HttpResponse('ok')

soluong=''

def index_1(request):
    return render(request, 'QRcodes/trang_chu_2.html')



def index(request):
    nhan_vien = Employee.objects.all()
   
    return render(request, 'QRFill/QRcodes/index.html',{'nhan_vien':nhan_vien})


serial_data = ''

def chon_don_1 (request):
    # if 's_khachhang' not in request.session:
    #     return redirect('customer:dang-nhap')
    return render(request, 'QRFill/QRcodes/chon.html')


def chon_don_2(request):
    # if 's_khachhang' not in request.session:
    #     return redirect('customer:dang-nhap')
    return render(request, 'QRFill/QRcodes/chon_don_2.html')

def don_1_nguoi(request):
    # if 's_khachhang' not in request.session:
    #     return redirect('customer:dang-nhap')
    hovaten_string=''
    so_cccd = ''
    so_cccd_cu =''
    ngay_cap_chuan=''
    gioi_tinh=''
    ngay_sinh_chuan=''
    gioi_tinh=''

    data_json_1=''

    if request.method == 'POST' or 'scan_data_1' in request.POST:
        thong_tin=request.POST.get('thong_tin')
     
        
        if thong_tin:
          
            thong_tin = thong_tin.rstrip()
            # cut "\r\n" at last of string
            thong_tin=thong_tin.replace('|',"*")
    
            
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            print(ho_va_ten)
            
        
            data_json_1 = {
            'so_cccd': so_cccd.replace('\x00',''),
            'so_cmnd_cu' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten':ho_va_ten,
            'ngay_sinh': ngay_sinh,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00','')
            
        }
        
        request.session['scan_data_1'] = data_json_1
        
        # Đường dẫn tới file Excel của bạn
        file_path = file_path = MEDIA_ROOT + 'QRcodes/reports/don_1_nguoi/data.xlsx'
        # Mở file Excel hiện có
        workbook = load_workbook(filename=file_path)
        # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
        sheet = workbook['Sheet1']
        # Dữ liệu mới bạn muốn thêm
        new_data = [so_cccd,so_cccd_cu,ho_va_ten,gioi_tinh,dia_chi,ngay_cap,thanh_pho_chuan,thoi_gian ]
        # Thêm dữ liệu vào hàng mới trong sheet
        sheet.append(new_data)
        # Lưu lại file Excel sau khi thêm dữ liệu
        workbook.save(filename=file_path)



            
    return render(request, 'QRcodes/don_1_nguoi.html', {'data_1':data_json_1})
    



@csrf_exempt

def don_2_nguoi(request):
    ho_va_ten=''
    ho_va_ten_2=''
    so_cccd_2 = ''
    so_cccd_cu_2 =''
    gioi_tinh_2=''
    dia_chi=''
    dia_chi_2=''
    gioi_tinh_2=''
    so_cccd = ''
    so_cccd_cu =''
    gioi_tinh=''
    dia_chi_2=''
    gioi_tinh=''
    ngay_sinh=''
    ngay_sinh_2=''
    ngay_cap=''
    ngay_cap_2=''
    thoi_gian=''
    data_json_1=''
    data_json_2=''

    if request.method == 'POST' and 'scan_data_1' in request.POST:
        thong_tin=request.POST.get('thong_tin')     
        if thong_tin:
          
            thong_tin = thong_tin.rstrip()
            # cut "\r\n" at last of string
            thong_tin=thong_tin.replace('|',"*")
    
            
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
    
            
        
            data_json_1 = {
            'so_cccd': so_cccd.replace('\x00',''),
            'so_cmnd_cu' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten':ho_va_ten,
            'ngay_sinh': ngay_sinh,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00','')
            
        }
        request.session['scan_data_1'] = data_json_1    
            
    if request.method == 'POST' and 'scan_data_2' in request.POST:
        thong_tin_2=request.POST.get('thong_tin_2')
        
        if thong_tin_2:
          
            thong_tin_2 = thong_tin_2.rstrip()
            # cut "\r\n" at last of string
            thong_tin_2=thong_tin_2.replace('|',"*")
    
            
            thong_tin_2=thong_tin_2.split('*')
            so_cccd_2 = thong_tin_2[0]
            ho_va_ten_2 = thong_tin_2[2].upper()
            ngay_sinh_2 = thong_tin_2[3]
            ngay_sinh_2= datetime.strptime(ngay_sinh_2, "%d%m%Y")
            ngay_sinh_2 = datetime.strftime(ngay_sinh_2, "%d/%m/%Y")
            
            gioi_tinh_2 = thong_tin_2[4]
            dia_chi_2 = thong_tin_2[5]
            ngay_cap_2 = thong_tin_2[6]
            ngay_cap_2= datetime.strptime(ngay_cap_2, "%d%m%Y")
            ngay_cap_2 = datetime.strftime(ngay_cap_2, "%d/%m/%Y")
            thanh_pho_2 = dia_chi_2.replace('\x00','').split(',')
            thanh_pho_chuan_2=thanh_pho_2[-1]
            thoi_gian_2 = datetime.now()
            thoi_gian_2= thoi_gian_2.strftime("%d/%m/%Y %H:%M")

        
        
            data_json_2 = {
            'so_cccd_2': so_cccd_2.replace('\x00',''),
            'so_cmnd_cu_2' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten_2':ho_va_ten_2,
            'ngay_sinh_2': ngay_sinh_2,
            'gioi_tinh_2':gioi_tinh_2.replace('\x00',''),
            'dia_chi_2': dia_chi_2.replace('\x00',''),
            'ngay_cap_2':ngay_cap_2.replace('\x00',''),
            'thanh_pho_2': thanh_pho_chuan_2.replace('\x00','')
            
        }
        request.session['scan_data_2'] = data_json_2
         # Đường dẫn tới file Excel của bạn
    data_1 = request.session.get('scan_data_1')
    data_2 =data_json_2
    if data_1 and data_2:
        thoi_gian = datetime.now()
        thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
        
        data_1.update(data_2)   
        print(data_1)
        c= type(data_1)
        print(c)
        file_path = MEDIA_ROOT + 'QRcodes/reports/don_2_nguoi/2_cccd.xlsx'

        # Mở file Excel hiện có
        workbook = load_workbook(filename=file_path)

        # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
        sheet = workbook['Sheet1']

        # Dữ liệu mới bạn muốn thêm
        new_data = [data_1["so_cccd"],
                    data_1["ho_va_ten"],
                    data_1["ngay_sinh"],
                    data_1["gioi_tinh"],
                    data_1["dia_chi"],
                    data_1["ngay_cap"],
                    data_1["so_cccd_2"],
                    data_1["ho_va_ten_2"],
                    data_1["ngay_sinh_2"],
                    data_1["gioi_tinh_2"],
                    data_1["dia_chi_2"],
                    data_1["ngay_cap_2"],
                    thoi_gian ]
        
        # Thêm dữ liệu vào hàng mới trong sheet
        sheet.append(new_data)

        # Lưu lại file Excel sau khi thêm dữ liệu
        workbook.save(filename=file_path)



    return render(request,'QRFill/QRcodes/don_2_nguoi.html', {'data_1':data_json_1,
                                                        'data_2': data_json_2})

def ket_hon(request):
    ho_va_ten=''
    ho_va_ten_2=''
    so_cccd_2 = ''
    so_cccd_cu_2 =''
    gioi_tinh_2=''
    dia_chi=''
    dia_chi_2=''
    gioi_tinh_2=''
    so_cccd = ''
    so_cccd_cu =''
    gioi_tinh=''
    dia_chi_2=''
    gioi_tinh=''
    ngay_sinh=''
    ngay_sinh_2=''
    ngay_cap=''
    ngay_cap_2=''
    thoi_gian=''
    data_json_1=''
    data_json_2=''

    
    if request.method == 'POST' and 'scan_data_1' in request.POST:
        thong_tin=request.POST.get('thong_tin')
        print(thong_tin)
        
        if thong_tin:
          
            thong_tin = thong_tin.rstrip()
            # cut "\r\n" at last of string
            thong_tin=thong_tin.replace('|',"*")
            print(thong_tin)
            
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            print(ho_va_ten)
            
        
            data_json_1 = {
            'so_cccd': so_cccd.replace('\x00',''),
            'so_cmnd_cu' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten':ho_va_ten,
            'ngay_sinh': ngay_sinh,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00','')
            
        }
           
        request.session['scan_data_1'] = data_json_1
        
 
            
    if request.method == 'POST' and 'scan_data_2' in request.POST:
        thong_tin_2=request.POST.get('thong_tin_2')
        
        if thong_tin_2:
          
            thong_tin_2 = thong_tin_2.rstrip()
            # cut "\r\n" at last of string
            thong_tin_2=thong_tin_2.replace('|',"*")
            print(thong_tin_2)
            
            thong_tin_2=thong_tin_2.split('*')
            so_cccd_2 = thong_tin_2[0]
            ho_va_ten_2 = thong_tin_2[2].upper()
            ngay_sinh_2 = thong_tin_2[3]
            ngay_sinh_2= datetime.strptime(ngay_sinh_2, "%d%m%Y")
            ngay_sinh_2 = datetime.strftime(ngay_sinh_2, "%d/%m/%Y")
            
            gioi_tinh_2 = thong_tin_2[4]
            dia_chi_2 = thong_tin_2[5]
            ngay_cap_2 = thong_tin_2[6]
            ngay_cap_2= datetime.strptime(ngay_cap_2, "%d%m%Y")
            ngay_cap_2 = datetime.strftime(ngay_cap_2, "%d/%m/%Y")
            thanh_pho_2 = dia_chi_2.replace('\x00','').split(',')
            thanh_pho_chuan_2=thanh_pho_2[-1]
            thoi_gian_2 = datetime.now()
            thoi_gian_2= thoi_gian_2.strftime("%d/%m/%Y %H:%M")
            print(thanh_pho_chuan_2)
        
        
            data_json_2 = {
            'so_cccd_2': so_cccd_2.replace('\x00',''),
            'so_cmnd_cu_2' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten_2':ho_va_ten_2,
            'ngay_sinh_2': ngay_sinh_2,
            'gioi_tinh_2':gioi_tinh_2.replace('\x00',''),
            'dia_chi_2': dia_chi_2.replace('\x00',''),
            'ngay_cap_2':ngay_cap_2.replace('\x00',''),
            'thanh_pho_2': thanh_pho_chuan_2.replace('\x00','')
            
        }
           
        request.session['scan_data_2'] = data_json_2
        
    data_1 = request.session.get('scan_data_1')
    data_2 =data_json_2
    if data_1 and data_2:
        thoi_gian = datetime.now()
        thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
        data_1.update(data_2)   
        print(data_1)
        c= type(data_1)
        print(c)
        file_path = MEDIA_ROOT + 'QRcodes/reports/don_2_nguoi/ket_hon.xlsx'

        # Mở file Excel hiện có
        workbook = load_workbook(filename=file_path)

        # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
        sheet = workbook['Sheet1']

        # Dữ liệu mới bạn muốn thêm
        new_data = [data_1["so_cccd"],
                    data_1["ho_va_ten"],
                    data_1["ngay_sinh"],
                    data_1["gioi_tinh"],
                    data_1["dia_chi"],
                    data_1["ngay_cap"],
                    data_1["so_cccd_2"],
                    data_1["ho_va_ten_2"],
                    data_1["ngay_sinh_2"],
                    data_1["gioi_tinh_2"],
                    data_1["dia_chi_2"],
                    data_1["ngay_cap_2"],
                    thoi_gian ]
        print(new_data)
        
        # Thêm dữ liệu vào hàng mới trong sheet
        sheet.append(new_data)

        # Lưu lại file Excel sau khi thêm dữ liệu
        workbook.save(filename=file_path)

        # Luu File
        
    
 

    return render(request,'QRFill/QRcodes/dien_ket_hon.html', {'data_1':data_json_1,
                                                    'data_2': data_json_2})


def don_khai_tu(request):
    ho_va_ten=''
    ho_va_ten_2=''
    so_cccd_2 = ''
    so_cccd_cu_2 =''
    gioi_tinh_2=''
    dia_chi=''
    dia_chi_2=''
    gioi_tinh_2=''
    so_cccd = ''
    so_cccd_cu =''
    gioi_tinh=''
    dia_chi_2=''
    gioi_tinh=''
    ngay_sinh=''
    ngay_sinh_2=''
    ngay_cap=''
    ngay_cap_2=''
    thoi_gian=''
    data_json_1=''
    data_json_2=''
    if request.method == 'POST' and 'scan_data_1' in request.POST:
        thong_tin=request.POST.get('thong_tin')
        print(thong_tin)
        
        if thong_tin:
          
            thong_tin = thong_tin.rstrip()
            # cut "\r\n" at last of string
            thong_tin=thong_tin.replace('|',"*")
            print(thong_tin)
            
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            print(ngay_sinh)
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            print(ngay_sinh)
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            print(ho_va_ten)
            
        
            data_json_1 = {
            'so_cccd': so_cccd.replace('\x00',''),
            'so_cmnd_cu' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten':ho_va_ten,
            'ngay_sinh': ngay_sinh,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00','')
            
        }
            request.session['scan_data_1'] = data_json_1
           
        data_1 = request.session.get('scan_data_1')  
                 
    if request.method == 'POST' and 'scan_data_2' in request.POST:
        thong_tin_2=request.POST.get('thong_tin_2')
        
        if thong_tin_2:
          
            thong_tin_2 = thong_tin_2.rstrip()
            # cut "\r\n" at last of string
            thong_tin_2=thong_tin_2.replace('|',"*")
            print(thong_tin_2)
            
            thong_tin_2=thong_tin_2.split('*')
            so_cccd_2 = thong_tin_2[0]
            ho_va_ten_2 = thong_tin_2[2].upper()
            ngay_sinh_2 = thong_tin_2[3]
            ngay_sinh_2= datetime.strptime(ngay_sinh_2, "%d%m%Y")
            ngay_sinh_2 = datetime.strftime(ngay_sinh_2, "%d/%m/%Y")
            
            gioi_tinh_2 = thong_tin_2[4]
            dia_chi_2 = thong_tin_2[5]
            ngay_cap_2 = thong_tin_2[6]
            ngay_cap_2= datetime.strptime(ngay_cap_2, "%d%m%Y")
            ngay_cap_2 = datetime.strftime(ngay_cap_2, "%d/%m/%Y")
            thanh_pho_2 = dia_chi_2.replace('\x00','').split(',')
            thanh_pho_chuan_2=thanh_pho_2[-1]
            thoi_gian_2 = datetime.now()
            thoi_gian_2= thoi_gian_2.strftime("%d/%m/%Y %H:%M")
            print(thanh_pho_chuan_2)
        
        
            data_json_2 = {
            'so_cccd_2': so_cccd_2.replace('\x00',''),
            'so_cmnd_cu_2' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten_2':ho_va_ten_2,
            'ngay_sinh_2': ngay_sinh_2,
            'gioi_tinh_2':gioi_tinh_2.replace('\x00',''),
            'dia_chi_2': dia_chi_2.replace('\x00',''),
            'ngay_cap_2':ngay_cap_2.replace('\x00',''),
            'thanh_pho_2': thanh_pho_chuan_2.replace('\x00','')
            
            
        }
            request.session['scan_data_2'] = data_json_2
    
    data_1 = request.session.get('scan_data_1')
    data_2 =data_json_2
    if data_1 and data_2:
        thoi_gian = datetime.now()
        thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
        data_1.update(data_2)   
        print(data_1)
        file_path = MEDIA_ROOT + 'QRcodes/reports/don_2_nguoi/khai_tu.xlsx'

        # Mở file Excel hiện có
        workbook = load_workbook(filename=file_path)

        # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
        sheet = workbook['Sheet1']

        # Dữ liệu mới bạn muốn thêm
        new_data = [data_1["so_cccd"],
                    data_1["ho_va_ten"],
                    data_1["ngay_sinh"],
                    data_1["gioi_tinh"],
                    data_1["dia_chi"],
                    data_1["ngay_cap"],
                    data_1["so_cccd_2"],
                    data_1["ho_va_ten_2"],
                    data_1["ngay_sinh_2"],
                    data_1["gioi_tinh_2"],
                    data_1["dia_chi_2"],
                    data_1["ngay_cap_2"],
                    thoi_gian ]
        print(new_data)
        
        # Thêm dữ liệu vào hàng mới trong sheet
        sheet.append(new_data)

        # Lưu lại file Excel sau khi thêm dữ liệu
        workbook.save(filename=file_path)

        # Luu File
    
    return render(request,'QRFill/QRcodes/khai_tu.html', {'data_1':data_json_1,
                                                    'data_2': data_json_2})



def khai_tu_1(request):
    x = datetime.now()
    ngay=x.day
    thang=x.month
    nam=x.year
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')
    
    
    if data_1 and data_2:
     
        ho_va_ten=data_1['ho_va_ten']
        ngay_sinh=data_1['ngay_sinh']
        so_cccd=data_1['so_cccd']
            
        # 
        ho_va_ten_2=data_2['ho_va_ten_2']
        ngay_sinh_2=data_2['ngay_sinh_2']
        so_cccd_2=data_2['so_cccd_2']
        
     
    if request.POST.get('in'):
       
         
        if data_1:
            del request.session['scan_data_1']
        if data_2:
            del request.session['scan_data_2']
        return render(request, 'QRcodes/result.html')
        
    
   
    
    return render(request,'QRFill/QRcodes/khai_tu/khai_tu_1.html',{ 
        'data1': data_1,
        'ngay':str(ngay),
        'thang':str(thang),
        'nam':str(nam),})
    
  
  
def ket_hon_1(request):
    x = datetime.now()
    ngay=x.day
    thang=x.month
    nam=x.year
    soluong=''
    
    
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')

    
    if data_1 and data_2:
     
        ho_va_ten=data_1['ho_va_ten']
        ngay_sinh=data_1['ngay_sinh']
        so_cccd=data_1['so_cccd']
      
        
        # 
        ho_va_ten_2=data_2['ho_va_ten_2']
        ngay_sinh_2=data_2['ngay_sinh_2']
        so_cccd_2=data_2['so_cccd_2']
        
    if request.POST.get('in'):
      
   
        if data_1:
            del request.session['scan_data_1']
        if data_2:
            del request.session['scan_data_2']
        
        
        # return redirect(url)
        return render(request, 'QRcodes/result.html')
        
     
    
    return render(request,'QRFill/QRcodes/ket_hon/ket_hon.html',{
        'ngay':ngay,
        'thang':thang,
        'nam':nam,
        'so_luong' : soluong,
        
    })
    
   
def trich_luc_ho_tich_1(request):
    x = datetime.now()
    ngay=x.day
    thang=x.month
    nam=x.year
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')
 
  
    if request.POST.get('in'):
        loai_don = request.POST.get('loai_don')
        so_luong = request.POST.get('so_luong')      
         
        if data_1:
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y")
            file_path = MEDIA_ROOT + 'QRcodes/reports/data/trich_luc_ho_tich/trich_luc_ho_tich.xlsx'

            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)

            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']

            # Dữ liệu mới bạn muốn thêm
            new_data = [
                        '',
                        thoi_gian,
                        f'{data_1["ho_va_ten"]}, Số CCCD {data_1["so_cccd"]}',
                        loai_don,
                        '',
                        so_luong
    
                     
                        ]
            
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)

            # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)
            del request.session['scan_data_1']
        if data_2:
            del request.session['scan_data_2']
    # Sử dụng reverse() để tạo URL với id và view name
        
        
        # return redirect(url)
        return render(request, 'QRcodes/result.html')
   
    return render(request, 'QRFill/QRcodes/ho_tich/ho_tich_1.html',{
        'ngay':ngay,
        'thang':thang,
        'nam':nam,
    })
    
 
    
def trich_luc_ho_tich_2(request):
    x = datetime.now()
    ngay=x.day
    thang=x.month
    nam=x.year
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')
    
        
    if request.POST.get('in'):
        
         
        if data_1:
            loai_don = request.POST.get('loai_don')
            so_luong = request.POST.get('so_luong')
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y")
            file_path = MEDIA_ROOT + 'QRcodes/reports/data/trich_luc_ho_tich/trich_luc_ho_tich.xlsx'

            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)

            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']

            # Dữ liệu mới bạn muốn thêm
            new_data = [
                        '',
                        thoi_gian,
                        f'{data_1["ho_va_ten"]}, Số CCCD {data_1["so_cccd"]}',
                        loai_don,
                        '',
                        so_luong
    
                     
                        ]
            
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)

            # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)
            del request.session['scan_data_1']
        if data_2:
            del request.session['scan_data_2']
        
        
        # return redirect(url)
        return render(request, 'QRcodes/result.html')
    return render(request, 'QRFill/QRcodes/ho_tich/ho_tich_2.html',{
        'ngay':ngay,
        'thang':thang,
        'nam':nam,
    })
    

def xac_nhan_hon_nhan(request):
    x = datetime.now()
    ngay=x.day
    thang=x.month
    nam=x.year
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')

        
    if request.POST.get('in'):
        dan_toc_1 = request.POST.get('dan_toc_1')
        tinh_trang = request.POST.get('tinh_trang_ket_hon_1')
        muc_dich = request.POST.get('muc_dich')
        print(tinh_trang)
        data_1 = request.session.get('scan_data_1')
   
        if data_1 :
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            file_path = MEDIA_ROOT + 'QRcodes/reports/data/Xac_nhan_tinh_trang/xntthn_1.xlsx'

            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)

            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']

            # Dữ liệu mới bạn muốn thêm
            new_data = [
                        '',
                        '',
                        thoi_gian,
                        'Bản thân',
                        data_1["ho_va_ten"],
                        data_1["ngay_sinh"],
                        data_1["gioi_tinh"],
                        dan_toc_1,
                        'Việt Nam',
                        data_1["dia_chi"],
                        f' Căn cước công dân số {data_1["so_cccd"]} do Cục CS QLHC về trật tự xã hội cấp ngày {data_1["ngay_cap"]}',    
                        tinh_trang,
                        muc_dich,
                        ]
            
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)

            # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)

        

        if data_1:
            del request.session['scan_data_1']
        if data_2:
            del request.session['scan_data_2']
        
        # return redirect(url)
        return render(request, 'QRcodes/result.html')

    return render(request,'QRFill/QRcodes/hon_nhan/hon_nhan_1.html',{
        'ngay':ngay,
        'thang':thang,
        'nam':nam,})
    
    
def xac_nhan_hon_nhan_2(request):
    x = datetime.now()
    ngay=x.day
    thang=x.month
    nam=x.year
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')

    if data_1 and data_2:
     
        ho_va_ten=data_1['ho_va_ten']
        ngay_sinh=data_1['ngay_sinh']
        so_cccd=data_1['so_cccd']
      
        
        # 
        ho_va_ten_2=data_2['ho_va_ten_2']
        ngay_sinh_2=data_2['ngay_sinh_2']
        so_cccd_2=data_2['so_cccd_2']
        
  
    if request.POST.get('in'):
        dan_toc_1 = request.POST.get('dan_toc_1')
        tinh_trang = request.POST.get('tinh_trang_ket_hon_1')
        muc_dich = request.POST.get('muc_dich')
        quan_he = request.POST.get('quan_he')
        print(quan_he)
        print(tinh_trang)
        data_1 = request.session.get('scan_data_1')
        data_2 = request.session.get('scan_data_2')
        print(data_2)
   
        if data_1 and data_2 :
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            file_path = MEDIA_ROOT + 'QRcodes/reports/data/Xac_nhan_tinh_trang/xntthn_2.xlsx'

            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)

            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']

            # Dữ liệu mới bạn muốn thêm
            new_data = [
                        '',
                        '',
                        thoi_gian,  
                        data_1["ho_va_ten"],
                        data_1["ngay_sinh"],
                
                        data_1["dia_chi"],
                        f' Căn cước công dân số {data_1["so_cccd"]} do Cục CS QLHC về trật tự xã hội cấp ngày {data_1["ngay_cap"]}',
                        quan_he,
                        # 
                        data_2["ho_va_ten_2"],
                        data_2["ngay_sinh_2"],
                        data_2["gioi_tinh_2"],
                        dan_toc_1,
                        'Việt Nam',
                        data_1["dia_chi"],
                        f' Căn cước công dân số {data_1["so_cccd"]} do Cục CS QLHC về trật tự xã hội cấp ngày {data_1["ngay_cap"]}',
                        tinh_trang,
                        muc_dich,
                        ]
            
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)

            # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)

        

        if data_1:
            del request.session['scan_data_1']
        if data_2:
            del request.session['scan_data_2']
        
        # return redirect(url)
        return render(request, 'QRcodes/result.html')
    

    return render(request,'QRFill/QRcodes/hon_nhan/hon_nhan_2.html',{
        'ngay':ngay,
        'thang':thang,
        'nam':nam,})





#Khai_sinh


def don_khai_sinh(request):
    ho_va_ten=''
    ho_va_ten_2=''
    ho_va_ten_3=''
    so_cccd=''
    so_cccd_2 = ''
    so_cccd_3 = ''
    so_cccd_cu_2 =''
    so_cccd_cu_3 =''
    gioi_tinh_2=''
    gioi_tinh_3=''
    dia_chi=''
    dia_chi_2=''
    dia_chi_3=''
    gioi_tinh_2=''
    gioi_tinh_3=''
    so_cccd_cu =''
    gioi_tinh=''
    dia_chi_2=''
    gioi_tinh=''
    ngay_sinh=''
    ngay_sinh_2=''
    ngay_sinh_3=''
    ngay_cap=''
    ngay_cap_2=''
    ngay_cap_3=''
    thoi_gian=''
    data_json_1=''
    data_json_2=''
    data_json_3=''
    if request.method == 'POST' and 'scan_data_1' in request.POST:
        thong_tin=request.POST.get('thong_tin')
        print(thong_tin)
        
        if thong_tin:
          
            thong_tin = thong_tin.rstrip()
            # cut "\r\n" at last of string
            thong_tin=thong_tin.replace('|',"*")
            print(thong_tin)
            
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            print(ho_va_ten)
            
        
            data_json_1 = {
            'so_cccd': so_cccd.replace('\x00',''),
            'so_cmnd_cu' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten':ho_va_ten,
            'ngay_sinh': ngay_sinh,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00','')
            
            }
        request.session['scan_data_1'] = data_json_1
       
     
        
            
    if request.method == 'POST' and 'scan_data_2' in request.POST:
        thong_tin_2=request.POST.get('thong_tin_2')
        
        if thong_tin_2:
          
            thong_tin_2 = thong_tin_2.rstrip()
            # cut "\r\n" at last of string
            thong_tin_2=thong_tin_2.replace('|',"*")
            print(thong_tin_2)
            
            thong_tin_2=thong_tin_2.split('*')
            so_cccd_2 = thong_tin_2[0]
            ho_va_ten_2 = thong_tin_2[2].upper()
            ngay_sinh_2 = thong_tin_2[3]
            ngay_sinh_2= datetime.strptime(ngay_sinh_2, "%d%m%Y")
            ngay_sinh_2 = datetime.strftime(ngay_sinh_2, "%d/%m/%Y")
            
            gioi_tinh_2 = thong_tin_2[4]
            dia_chi_2 = thong_tin_2[5]
            ngay_cap_2 = thong_tin_2[6]
            ngay_cap_2= datetime.strptime(ngay_cap_2, "%d%m%Y")
            ngay_cap_2 = datetime.strftime(ngay_cap_2, "%d/%m/%Y")
            thanh_pho_2 = dia_chi_2.replace('\x00','').split(',')
            thanh_pho_chuan_2=thanh_pho_2[-1]
            thoi_gian_2 = datetime.now()
            thoi_gian_2= thoi_gian_2.strftime("%d/%m/%Y %H:%M")
            print(thanh_pho_chuan_2)
        
        
            data_json_2 = {
            'so_cccd_2': so_cccd_2.replace('\x00',''),
            'so_cmnd_cu_2' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten_2':ho_va_ten_2,
            'ngay_sinh_2': ngay_sinh_2,
            'gioi_tinh_2':gioi_tinh_2.replace('\x00',''),
            'dia_chi_2': dia_chi_2.replace('\x00',''),
            'ngay_cap_2':ngay_cap_2.replace('\x00',''),
            'thanh_pho_2': thanh_pho_chuan_2.replace('\x00','')
            
        }
  
        request.session['scan_data_2'] = data_json_2
        
    if request.method == 'POST' and 'scan_data_3' in request.POST:
        thong_tin_3=request.POST.get('thong_tin_3')
        
        if thong_tin_3:
          
            thong_tin_3 = thong_tin_3.rstrip()
            # cut "\r\n" at last of string
            thong_tin_3=thong_tin_3.replace('|',"*")
            print(thong_tin_3)
            
            thong_tin_3=thong_tin_3.split('*')
            so_cccd_3 = thong_tin_3[0]
            so_cccd_cu_3 = thong_tin_3[1]
            ho_va_ten_3 = thong_tin_3[2].upper()
            ngay_sinh_3 = thong_tin_3[3]
            ngay_sinh_3= datetime.strptime(ngay_sinh_3, "%d%m%Y")
            ngay_sinh_3 = datetime.strftime(ngay_sinh_3, "%d/%m/%Y")
            
            gioi_tinh_3 = thong_tin_3[4]
            dia_chi_3 = thong_tin_3[5]
            ngay_cap_3 = thong_tin_3[6]
            ngay_cap_3= datetime.strptime(ngay_cap_3, "%d%m%Y")
            ngay_cap_3 = datetime.strftime(ngay_cap_3, "%d/%m/%Y")
            thanh_pho_3 = dia_chi_3.replace('\x00','').split(',')
            thanh_pho_chuan_3=thanh_pho_3[-1]
            thoi_gian_3 = datetime.now()
            thoi_gian_3= thoi_gian_3.strftime("%d/%m/%Y %H:%M")
            print(thanh_pho_chuan_3)
        
        
            data_json_3 = {
            'so_cccd_3': so_cccd_3.replace('\x00',''),
            'so_cmnd_cu_3' : so_cccd_cu_3.replace('\x00',''),
            'ho_va_ten_3':ho_va_ten_3,
            'ngay_sinh_3': ngay_sinh_3,
            'gioi_tinh_3':gioi_tinh_3.replace('\x00',''),
            'dia_chi_3': dia_chi_3.replace('\x00',''),
            'ngay_cap_3':ngay_cap_3.replace('\x00',''),
            'thanh_pho_3': thanh_pho_chuan_3.replace('\x00','')
            
        }
 
        request.session['scan_data_3'] = data_json_3
        
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')
   
    data_3 =data_json_3
    if data_1 and data_2 and data_3:
        thoi_gian = datetime.now()
        thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
        
        data_1.update(data_2)
        data_1.update(data_3)   
        print(data_1)
        c= type(data_1)
        print(c)
        file_path = MEDIA_ROOT + 'QRcodes/reports/khai_sinh/khai_sinh.xlsx'

        # Mở file Excel hiện có
        workbook = load_workbook(filename=file_path)

        # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
        sheet = workbook['Sheet1']

        # Dữ liệu mới bạn muốn thêm
        new_data = [data_1["so_cccd"],
                    data_1["ho_va_ten"],
                    data_1["ngay_sinh"],
                    data_1["gioi_tinh"],
                    data_1["dia_chi"],
                    data_1["ngay_cap"],
                    data_1["so_cccd_2"],
                    data_1["ho_va_ten_2"],
                    data_1["ngay_sinh_2"],
                    data_1["gioi_tinh_2"],
                    data_1["dia_chi_2"],
                    data_1["ngay_cap_2"],
                    data_1["so_cccd_3"],
                    data_1["ho_va_ten_3"],
                    data_1["ngay_sinh_3"],
                    data_1["gioi_tinh_3"],
                    data_1["dia_chi_3"],
                    data_1["ngay_cap_3"],
                    thoi_gian ]
        
        # Thêm dữ liệu vào hàng mới trong sheet
        sheet.append(new_data)

        # Lưu lại file Excel sau khi thêm dữ liệu
        workbook.save(filename=file_path)

    
             
    
 

    return render(request,'QRFill/QRcodes/don_3_nguoi.html', {'data_1':data_json_1,
                                                    'data_2': data_json_2,
                                                      'data_3': data_json_3 } )


# luu--khai_sinb

def khai_sinh_1(request):
    name_string=''
    date_string=''
    date_text=''
    x = datetime.now()
    ngay=x.day
    thang=x.month
    nam=x.year
    soluong=''
    ho_va_ten=''
    so_cccd=''    # 
    ho_va_ten_me=''
    ngay_sinh_me=''
    
    ho_va_ten_cha=''
    ngay_sinh_cha=''
    

    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')
    data_3 = request.session.get('scan_data_3')
    

    if request.POST.get('ghi_1'):
        date_string=request.POST.get('ghi_2')
        quan_he = request.POST.get('quan_he')
        print(date_string)
        name_string=request.POST.get('ten_khai_sinh')
        print(name_string)

        parts = date_string.split('/')
        print(parts)
        day = int(parts[0])
        month = int(parts[1])
        year = int(parts[2])
        day_word = num2words(day, to='ordinal', lang='vi')
        month_word = num2words(month, lang='vi')
        year_word = num2words(year, lang='vi')
        date_text = f''' Ngày {day_word} tháng {month_word} năm {year_word}'''
        return render(request,'QRFill/QRcodes/khai_sinh/khai_sinh.html',{'date_text':date_text,
                                                         'name_string': name_string,
                                                         'date_string':date_string,
                                                         'quan_he' : quan_he,
                                                         'ngay':ngay,
        'thang':thang,
        'nam':nam,})
        
    elif request.POST.get('in'):
        
        if data_1:
            del request.session['scan_data_1']
        if data_2:
    
            del request.session['scan_data_2']
            
        if data_3:
           
            del request.session['scan_data_3']
            
        
        
        # return redirect(url)
        
        return render(request, 'QRcodes/result.html')
    
        
    
    return render(request,'QRFill/QRcodes/khai_sinh/khai_sinh.html',{
        'ngay':ngay,
        'thang':thang,
        'nam':nam,
        'so_luong' : soluong,
        'date_text': date_text,
        'name_string': name_string,
        'date_string':date_string
        
    })
        


def upload_file(request):
    path_report=''
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')
    data_3 = request.session.get('scan_data_3')
    if data_1 :
        ho_va_ten_1=data_1['ho_va_ten']
        gioi_tinh_1 = data_1['gioi_tinh']
        ngay_sinh_1=data_1['ngay_sinh']
        so_cccd_1=data_1['so_cccd']
        dia_chi_1 = data_1['dia_chi']
        ngay_cap_1 = data_1['ngay_cap']
        # thanh_pho_1=data_1['thanh_pho']
    if data_1 and data_2:  
        ho_va_ten_1=data_1['ho_va_ten']
        gioi_tinh_1 = data_1['gioi_tinh']
        ngay_sinh_1=data_1['ngay_sinh']
        
        so_cccd_1=data_1['so_cccd']
        dia_chi_1 = data_1['dia_chi']
        ngay_cap_1 = data_1['ngay_cap']
        thanh_pho_1=data_1['thanh_pho']     
        ho_va_ten_2=data_2['ho_va_ten_2']
        ngay_sinh_2=data_2['ngay_sinh_2']
        so_cccd_2=data_2['so_cccd_2']
        gioi_tinh_2 = data_2['gioi_tinh_2']
        ngay_sinh_2=data_2['ngay_sinh_2']
        so_cccd_2=data_2['so_cccd_2']
        dia_chi_2 = data_2['dia_chi_2']
        ngay_cap_2 = data_2['ngay_cap_2']
        thanh_pho_2=data_2['thanh_pho_2']
    if data_1 and data_2 and data_3:

        ho_va_ten_1=data_1['ho_va_ten']
        gioi_tinh_1 = data_1['gioi_tinh']
        ngay_sinh_1=data_1['ngay_sinh']
        so_cccd_1=data_1['so_cccd']
        dia_chi_1 = data_1['dia_chi']
        ngay_cap_1 = data_1['ngay_cap']
        thanh_pho_1=data_1['thanh_pho']     
        ho_va_ten_2=data_2['ho_va_ten_2']
        ngay_sinh_2=data_2['ngay_sinh_2']
        so_cccd_2=data_2['so_cccd_2']
        gioi_tinh_2 = data_2['gioi_tinh_2']
        ngay_sinh_2=data_2['ngay_sinh_2']
        so_cccd_2=data_2['so_cccd_2']
        dia_chi_2 = data_2['dia_chi_2']
        ngay_cap_2 = data_2['ngay_cap_2']
        thanh_pho_2=data_2['thanh_pho_2']
        ho_va_ten_3=data_3['ho_va_ten_3']
        gioi_tinh_3 = data_3['gioi_tinh_3']
        ngay_sinh_3=data_3['ngay_sinh_3']
        so_cccd_3=data_3['so_cccd_3']
        dia_chi_3 = data_3['dia_chi_3']
        ngay_cap_3 = data_3['ngay_cap_3']
        thanh_pho_3=data_3['thanh_pho_3']
        
        
    if request.method == "POST" and 'upload_file_1' in request.POST:
        uploaded_file = request.FILES.get('upload_file_1')
        pythoncom.CoInitialize()
        word_app = win32.gencache.EnsureDispatch("Word.Application")
        
        # Lưu tệp đã tải lên vào đĩa cứng
        uploaded_file_path = os.path.join(MEDIA_ROOT + 'QRcodes\\reports\\', uploaded_file.name)
        with open(uploaded_file_path, 'wb') as destination_file:
            for chunk in uploaded_file.chunks():
                destination_file.write(chunk)

        # Mở tệp từ đường dẫn đã lưu
        doc = word_app.Documents.Open(uploaded_file_path)

        # Tiếp tục với phần còn lại của mã của bạn...
        file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr_1.xlsx'
        doc.MailMerge.OpenDataSource(file_path)
        
     
        # Tiếp tục với phần còn lại của mã của bạn...
        file_name = f' moi_{uploaded_file.name}'

        folder_report = folder_report = MEDIA_ROOT + 'QRcodes/reports/file_moi/'
        path_report = folder_report + file_name
        
        doc.SaveAs(path_report)
        doc.Close()
        
        # Đóng ứng dụng Word
        word_app.Quit()
        subprocess.Popen(["start", "winword",   path_report], shell=True)
        return render(request, 'QRcodes/a.html')
    
        
    if request.method == 'POST' and 'upload_file' in request.POST:
        uploaded_file = request.FILES['upload_file']
        print(uploaded_file)
        doc = Document(uploaded_file),
        document = MailMerge(uploaded_file)
        document.merge_fonts = {
        'eastAsia': 'Arial Unicode MS',
        'cs': 'Times New Roman',
        'hAnsi': 'Tahoma'   
    }
        if data_1 and not data_2:
            document.merge (
            ho_va_ten_1=data_1['ho_va_ten'],
            gioi_tinh_1 = data_1['gioi_tinh'],
            ngay_sinh_1=data_1['ngay_sinh'],
            so_cccd_1=data_1['so_cccd'],
            dia_chi_1 = data_1['dia_chi'],
            ngay_cap_1 = data_1['ngay_cap'],
            thanh_pho_1=data_1['thanh_pho'],)
            current_date = datetime.now() 
            file_name = ho_va_ten_1 + '_'  + current_date.strftime('%Y%m%d_%H%M%S') + '.docx'
            folder_report = MEDIA_ROOT + 'QRcodes\\reports\\'
            path_report = folder_report + file_name
            document.write(path_report)
            del request.session['scan_data_1'] 
            subprocess.Popen(["start", "winword",   path_report], shell=True)
                       
    
        if data_1 and data_2 and not data_3:
            document.merge (
            ho_va_ten_1=data_1['ho_va_ten'],
            gioi_tinh_1 = data_1['gioi_tinh'],
            ngay_sinh_1=data_1['ngay_sinh'],
            so_cccd_1=data_1['so_cccd'],
            dia_chi_1 = data_1['dia_chi'],
            ngay_cap_1 = data_1['ngay_cap'],
            thanh_pho_1=data_1['thanh_pho'],
            ho_va_ten_2=data_2['ho_va_ten_2'],
            gioi_tinh_2 = data_2['gioi_tinh_2'],
            ngay_sinh_2=data_2['ngay_sinh_2'],
            so_cccd_2=data_2['so_cccd_2'],
            dia_chi_2 = data_2['dia_chi_2'],
            ngay_cap_2 = data_2['ngay_cap_2'],
            thanh_pho_2=data_2['thanh_pho_2'],
            )
            current_date = datetime.now() 

            file_name = ho_va_ten_1 + '_'  + current_date.strftime('%Y%m%d_%H%M%S') + '.docx'
            folder_report = MEDIA_ROOT + 'QRcodes\\reports\\'
            path_report = folder_report + file_name
            document.write(path_report)
            subprocess.Popen(["start", "winword",   path_report], shell=True)
            del request.session['scan_data_1']
            del request.session['scan_data_2']
            
        if data_1 and data_2 and data_3:  
            document.merge (
            ho_va_ten_1=data_1['ho_va_ten'],
            gioi_tinh_1 = data_1['gioi_tinh'],
            ngay_sinh_1=data_1['ngay_sinh'],
            so_cccd_1=data_1['so_cccd'],
            dia_chi_1 = data_1['dia_chi'],
            ngay_cap_1 = data_1['ngay_cap'],
            thanh_pho_1=data_1['thanh_pho'],
            ho_va_ten_2=data_2['ho_va_ten_2'],
            gioi_tinh_2 = data_2['gioi_tinh_2'],
            ngay_sinh_2=data_2['ngay_sinh_2'],
            so_cccd_2=data_2['so_cccd_2'],
            dia_chi_2 = data_2['dia_chi_2'],
            ngay_cap_2 = data_2['ngay_cap_2'],
            thanh_pho_2=data_2['thanh_pho_2'],
        
            ho_va_ten_3=data_3['ho_va_ten_3'],
            gioi_tinh_3 = data_3['gioi_tinh_3'],
            ngay_sinh_3=data_3['ngay_sinh_3'],
            so_cccd_3=data_3['so_cccd_3'],
            dia_chi_3 = data_3['dia_chi_3'],
            ngay_cap_3 = data_3['ngay_cap_3'],
            thanh_pho_3=data_3['thanh_pho_3'],
        )
            current_date = datetime.now() 
            file_name = ho_va_ten_1 + '_'  + current_date.strftime('%Y%m%d_%H%M%S') + '.docx'
            folder_report = MEDIA_ROOT + 'QRcodes\\reports\\'
            path_report = folder_report + file_name
            document.write(path_report)
            subprocess.Popen(["start", "winword",   path_report], shell=True)

            del request.session['scan_data_1']
            del request.session['scan_data_2']
            del request.session['scan_data_3']
    
        # Xử lý và điền dữ liệu vào tệp


    return render(request, 'QRcodes/a.html')
  

def run_word_macro(request):

    if request.method == 'POST':
        uploaded_file = request.FILES['upload_file_1']
        pythoncom.CoInitialize()
        word_app = win32.gencache.EnsureDispatch("Word.Application")
        
        # Lưu tệp đã tải lên vào đĩa cứng
        uploaded_file_path = os.path.join(MEDIA_ROOT + 'QRcodes\\reports\\', uploaded_file.name)
        with open(uploaded_file_path, 'wb') as destination_file:
            for chunk in uploaded_file.chunks():
                destination_file.write(chunk)

        # Mở tệp từ đường dẫn đã lưu
        doc = word_app.Documents.Open(uploaded_file_path)

        # Tiếp tục với phần còn lại của mã của bạn...
        file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr_1.xlsx'
        doc.MailMerge.OpenDataSource(file_path)
        
     
        # Tiếp tục với phần còn lại của mã của bạn...
        file_name = f' moi_{uploaded_file.name}'

        folder_report = folder_report = MEDIA_ROOT + 'QRcodes/reports/file_moi/'
        path_report = folder_report + file_name
        
        doc.SaveAs(path_report)
        doc.Close()
        
        # Đóng ứng dụng Word
        word_app.Quit()
        subprocess.Popen(["start", "winword",   path_report], shell=True)
        
        
    return render(request, 'QRcodes/b.html')
# Đường dẫn đến tệp Excel
     

# Chạy macro và thực hiện mail merg





def quet_3_nguoi(request):
    ho_va_ten=''
    ho_va_ten_2=''
    ho_va_ten_3=''
    so_cccd=''
    so_cccd_2 = ''
    so_cccd_3 = ''
    so_cccd_cu_2 =''
    so_cccd_cu_3 =''
    gioi_tinh_2=''
    gioi_tinh_3=''
    dia_chi=''
    dia_chi_2=''
    dia_chi_3=''
    gioi_tinh_2=''
    gioi_tinh_3=''
    so_cccd_cu =''
    gioi_tinh=''
    dia_chi_2=''
    gioi_tinh=''
    ngay_sinh=''
    ngay_sinh_2=''
    ngay_sinh_3=''
    ngay_cap=''
    ngay_cap_2=''
    ngay_cap_3=''
    thoi_gian=''
    data_json_1=''
    data_json_2=''
    data_json_3=''
    if request.method == 'POST' and 'scan_data_1' in request.POST:
        thong_tin=request.POST.get('thong_tin')
        
        if thong_tin:
          
            thong_tin = thong_tin.rstrip()
            # cut "\r\n" at last of string
            thong_tin=thong_tin.replace('|',"*")
            
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            
        
            data_json_1 = {
            'so_cccd': so_cccd.replace('\x00',''),
            'so_cmnd_cu' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten':ho_va_ten,
            'ngay_sinh': ngay_sinh,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00','')
            
            }
            
        request.session['scan_data_1'] = data_json_1
                 
    if request.method == 'POST' and 'scan_data_2' in request.POST:
        thong_tin_2=request.POST.get('thong_tin_2')
        
        if thong_tin_2:
          
            thong_tin_2 = thong_tin_2.rstrip()
            # cut "\r\n" at last of string
            thong_tin_2=thong_tin_2.replace('|',"*")
            
            thong_tin_2=thong_tin_2.split('*')
            so_cccd_2 = thong_tin_2[0]
            ho_va_ten_2 = thong_tin_2[2].upper()
            ngay_sinh_2 = thong_tin_2[3]
            ngay_sinh_2= datetime.strptime(ngay_sinh_2, "%d%m%Y")
            ngay_sinh_2 = datetime.strftime(ngay_sinh_2, "%d/%m/%Y")
            
            gioi_tinh_2 = thong_tin_2[4]
            dia_chi_2 = thong_tin_2[5]
            ngay_cap_2 = thong_tin_2[6]
            ngay_cap_2= datetime.strptime(ngay_cap_2, "%d%m%Y")
            ngay_cap_2 = datetime.strftime(ngay_cap_2, "%d/%m/%Y")
            thanh_pho_2 = dia_chi_2.replace('\x00','').split(',')
            thanh_pho_chuan_2=thanh_pho_2[-1]
            thoi_gian_2 = datetime.now()
            thoi_gian_2= thoi_gian_2.strftime("%d/%m/%Y %H:%M")
        
        
            data_json_2 = {
            'so_cccd_2': so_cccd_2.replace('\x00',''),
            'so_cmnd_cu_2' : so_cccd_cu.replace('\x00',''),
            'ho_va_ten_2':ho_va_ten_2,
            'ngay_sinh_2': ngay_sinh_2,
            'gioi_tinh_2':gioi_tinh_2.replace('\x00',''),
            'dia_chi_2': dia_chi_2.replace('\x00',''),
            'ngay_cap_2':ngay_cap_2.replace('\x00',''),
            'thanh_pho_2': thanh_pho_chuan_2.replace('\x00','')
            
        }
  
        request.session['scan_data_2'] = data_json_2
        
    if request.method == 'POST' and 'scan_data_3' in request.POST:
        thong_tin_3=request.POST.get('thong_tin_3')
        
        if thong_tin_3:
          
            thong_tin_3 = thong_tin_3.rstrip()
            # cut "\r\n" at last of string
            thong_tin_3=thong_tin_3.replace('|',"*")
            
            thong_tin_3=thong_tin_3.split('*')
            so_cccd_3 = thong_tin_3[0]
            so_cccd_cu_3 = thong_tin_3[1]
            ho_va_ten_3 = thong_tin_3[2].upper()
            ngay_sinh_3 = thong_tin_3[3]
            ngay_sinh_3= datetime.strptime(ngay_sinh_3, "%d%m%Y")
            ngay_sinh_3 = datetime.strftime(ngay_sinh_3, "%d/%m/%Y")
            
            gioi_tinh_3 = thong_tin_3[4]
            dia_chi_3 = thong_tin_3[5]
            ngay_cap_3 = thong_tin_3[6]
            ngay_cap_3= datetime.strptime(ngay_cap_3, "%d%m%Y")
            ngay_cap_3 = datetime.strftime(ngay_cap_3, "%d/%m/%Y")
            thanh_pho_3 = dia_chi_3.replace('\x00','').split(',')
            thanh_pho_chuan_3=thanh_pho_3[-1]
            thoi_gian_3 = datetime.now()
            thoi_gian_3= thoi_gian_3.strftime("%d/%m/%Y %H:%M")

        
        
            data_json_3 = {
            'so_cccd_3': so_cccd_3.replace('\x00',''),
            'so_cmnd_cu_3' : so_cccd_cu_3.replace('\x00',''),
            'ho_va_ten_3':ho_va_ten_3,
            'ngay_sinh_3': ngay_sinh_3,
            'gioi_tinh_3':gioi_tinh_3.replace('\x00',''),
            'dia_chi_3': dia_chi_3.replace('\x00',''),
            'ngay_cap_3':ngay_cap_3.replace('\x00',''),
            'thanh_pho_3': thanh_pho_chuan_3.replace('\x00','')
            
        }
 
        request.session['scan_data_3'] = data_json_3
        
   
    data_1 = request.session.get('scan_data_1')
    data_2 = request.session.get('scan_data_2')
    data_3 =request.session.get('scan_data_3')
    # Take the filename to check
    filename = MEDIA_ROOT + 'QRcodes/reports/data/data_qr.xlsx'    
    try:
        myfile = open(filename, "r+") 
    except IOError:
         return HttpResponse('Kiểm tra và đóng file excel đang mở')

        # restart the loop
    if request.method == 'POST' and 'xac_nhan' in request.POST:
        
        if data_1 and not data_2: 
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr.xlsx'
            
            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)
            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']
            # Dữ liệu mới bạn muốn thêm
            new_data = [thoi_gian ,
                        data_1["so_cccd"],       
                        data_1["ho_va_ten"],
                        data_1["ngay_sinh"],
                        data_1["gioi_tinh"],
                        data_1["dia_chi"],
                        data_1["ngay_cap"],
                    
                        thoi_gian ]    
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)
         # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)
         
            workbook = Workbook()
            sheet = workbook.active
            data = [
                
                {   'thoi_gian' : thoi_gian,
                    'so_cccd_1' : data_1 ["so_cccd"],
                    'ho_va_ten_1' :data_1["ho_va_ten"],
                    'ngay_sinh_1': data_1["ngay_sinh"],
                    'gioi_tinh_1':data_1["gioi_tinh"],
                    'dia_chi_1':data_1["dia_chi"],
                    'ngay_cap': data_1["ngay_cap"]
                 }
            ]
            header = list(data[0].keys())
            sheet.append(header)

            # Sau đó, chèn dữ liệu từ từng từ điển
            for entry in data:
                row_data = [entry[key] for key in header]
                sheet.append(row_data)

            # Lưu tệp Excel mới
            excel_file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr_1.xlsx'
            workbook.save(excel_file_path)
            
          
            
            return redirect('qrcode:upload')

        if data_1 and data_2 and not data_3 :
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            
            data_1.update(data_2)

            file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr.xlsx'

            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)

            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']

            # Dữ liệu mới bạn muốn thêm
            new_data = [ 
                        thoi_gian,
                        data_1["so_cccd"],
                        data_1["ho_va_ten"],
                        data_1["ngay_sinh"],
                        data_1["gioi_tinh"],
                        data_1["dia_chi"],
                        data_1["ngay_cap"],
                        data_1["so_cccd_2"],
                        data_1["ho_va_ten_2"],
                        data_1["ngay_sinh_2"],
                        data_1["gioi_tinh_2"],
                        data_1["dia_chi_2"],
                        data_1["ngay_cap_2"],
                    
                         ]
            
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)

            # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)
            
            #### 
            
            workbook = Workbook()
            sheet = workbook.active
            data = [
                
                {   'thoi_gian' :thoi_gian,
                    'so_cccd_1' : data_1 ["so_cccd"],
                    'ho_va_ten_1' :data_1["ho_va_ten"],
                    'ngay_sinh_1': data_1["ngay_sinh"],
                    'gioi_tinh_1':data_1["gioi_tinh"],
                    'dia_chi_1':data_1["dia_chi"],
                    'ngay_cap_1': data_1["ngay_cap"],
                    'so_cccd_2' : data_1 ["so_cccd_2"],
                    'ho_va_ten_2' :data_1["ho_va_ten_2"],
                    'ngay_sinh_2': data_1["ngay_sinh_2"],
                    'gioi_tinh_2':data_1["gioi_tinh_2"],
                    'dia_chi_2':data_1["dia_chi_2"],
                    'ngay_cap_2': data_1["ngay_cap_2"],
                    
                 }
            ]
            header = list(data[0].keys())
            sheet.append(header)

            # Sau đó, chèn dữ liệu từ từng từ điển
            for entry in data:
                row_data = [entry[key] for key in header]
                sheet.append(row_data)

            # Lưu tệp Excel mới
            excel_file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr_1.xlsx'
            workbook.save(excel_file_path)
            ##
            return redirect('qrcode:upload')

        
        if data_1 and data_2 and data_3:
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            
            data_1.update(data_2)
            data_1.update(data_3)   
            c= type(data_1)

            file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr.xlsx'

            # Mở file Excel hiện có
            workbook = load_workbook(filename=file_path)

            # Chọn sheet bạn muốn thêm dữ liệu (ví dụ: 'Sheet1')
            sheet = workbook['Sheet1']

            # Dữ liệu mới bạn muốn thêm
            new_data = [
                         thoi_gian,
                        data_1["so_cccd"],
                        data_1["ho_va_ten"],
                        data_1["ngay_sinh"],
                        data_1["gioi_tinh"],
                        data_1["dia_chi"],
                        data_1["ngay_cap"],
                        data_1["so_cccd_2"],
                        data_1["ho_va_ten_2"],
                        data_1["ngay_sinh_2"],
                        data_1["gioi_tinh_2"],
                        data_1["dia_chi_2"],
                        data_1["ngay_cap_2"],
                        data_1["so_cccd_3"],
                        data_1["ho_va_ten_3"],
                        data_1["ngay_sinh_3"],
                        data_1["gioi_tinh_3"],
                        data_1["dia_chi_3"],
                        data_1["ngay_cap_3"],
                       ]
            
            # Thêm dữ liệu vào hàng mới trong sheet
            sheet.append(new_data)

            # Lưu lại file Excel sau khi thêm dữ liệu
            workbook.save(filename=file_path)
            
            
            workbook = Workbook()
            sheet = workbook.active
            data = [
                
                {   'thoi_gian' :thoi_gian,
                    'so_cccd_1' : data_1 ["so_cccd"],
                    'ho_va_ten_1' :data_1["ho_va_ten"],
                    'ngay_sinh_1': data_1["ngay_sinh"],
                    'gioi_tinh_1':data_1["gioi_tinh"],
                    'dia_chi_1':data_1["dia_chi"],
                    'ngay_cap_1': data_1["ngay_cap"],
                    'so_cccd_2' : data_1 ["so_cccd_2"],
                    'ho_va_ten_2' :data_1["ho_va_ten_2"],
                    'ngay_sinh_2': data_1["ngay_sinh_2"],
                    'gioi_tinh_2':data_1["gioi_tinh_2"],
                    'dia_chi_2':data_1["dia_chi_2"],
                    'ngay_cap_2': data_1["ngay_cap_2"],
                    'so_cccd_3' : data_1 ["so_cccd_3"],
                    'ho_va_ten_3' :data_1["ho_va_ten_3"],
                    'ngay_sinh_3': data_1["ngay_sinh_3"],
                    'gioi_tinh_3':data_1["gioi_tinh_3"],
                    'dia_chi_3':data_1["dia_chi_3"],
                    'ngay_cap_3': data_1["ngay_cap_3"],
                    
           
                    
                    
                 }
            ]
            header = list(data[0].keys())
            sheet.append(header)

            # Sau đó, chèn dữ liệu từ từng từ điển
            for entry in data:
                row_data = [entry[key] for key in header]
                sheet.append(row_data)

            # Lưu tệp Excel mới
            excel_file_path = MEDIA_ROOT + 'QRcodes/reports/data/data_qr_1.xlsx'
            workbook.save(excel_file_path)
            return redirect('qrcode:upload')



    return render(request,'QRcodes/quet_3_nguoi.html', {'data_1':data_json_1,
                                                    'data_2': data_json_2,
                                                      'data_3': data_json_3 } )


