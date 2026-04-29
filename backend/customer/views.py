from django.shortcuts import render, redirect, get_object_or_404
from backend.customer.models import *
from django.http import JsonResponse, HttpResponse
from collections import defaultdict
from django.urls import reverse
import time
from django.db.models import Max
from datetime import datetime
import qrcode
import qrcode.image.svg
from io import BytesIO
from django.conf import settings
from qrcode import make
from config.settings import MEDIA_ROOT
from django.views.decorators.csrf import csrf_exempt
from mailmerge import MailMerge
from docx import Document
from openpyxl import load_workbook
from num2words import num2words
from django.contrib.auth import authenticate, login
import subprocess
from openpyxl import Workbook
from django.db import connection
from django.utils import timezone
from django.db.models import Count
from django.core.exceptions import ObjectDoesNotExist
from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync
import os
import tempfile
import numpy as np

# Hardware layer (Modular Monolith)
from core.hardware.printer import print_ticket, convert_to_latin1
from core.hardware.audio import speak, save_tts, play_audio_file

# Business logic layer (services)
from backend.customer import services as customer_services

import json

# Đây là một ví dụ về một hàm cần một đối số, nhưng khi gọi lại nó, không truyền đối số.



def trang_chu_2(request):

    nhan_vien = Employee.objects.all()
    dich_vu = Service.objects.all()
    return render(request,'customer/trang_chu_lay_so.html',{'nhan_vien':nhan_vien,'dich_vu':dich_vu})
    
def dang_nhap(request):
    result_login=''
    if request.POST.get('btnDangNhap'):
    # Gán biến
        user_name = request.POST.get('user_name')
        mat_khau = request.POST.get('mat_khau')

        # Xử lý đọc thông tin từ CSDL
        nguoi_dung = Customer.objects.filter(user_name=user_name, password=mat_khau)
        if nguoi_dung.count() > 0:
            dict_nguoi_dung = nguoi_dung.values()[0]

            del(dict_nguoi_dung['password'])
            request.session['s_khachhang'] = dict_nguoi_dung
            return redirect('customer:trang-chu')
        else:
            result_login = '''
            <div class="alert alert-danger" role="alert">
                Đăng nhập thất bại. Vui lòng kiểm lại thông tin
            </div>
            '''

    return render(request, 'QRcodes/login.html', {
    
    'result_login': result_login,
   
    })
def chon_dich_vu(request):
    dich_vu= Service.objects.all()
    dich_vu_1_4= Service.objects.all()[:4]
    dich_vu_5_8 = Service.objects.filter(is_hidden=True)[4:8]
    nhan_vien=Employee.objects.all()
    services=Service.objects.all() 
    thoi_gian= Manage.objects.get(id=1)
    hinh_anh = Banner.objects.all()
    context=''
    for service in services:
        context = {'initial_service_status': service.is_active}

 
    return render(request,'customer/index_1.html',{'dich_vu':dich_vu, 'context':context,'nhan_vien':nhan_vien,'thoi_gian':thoi_gian,'hinh_anh':hinh_anh,'dich_vu_1_4':dich_vu_1_4,'dich_vu_5_8':dich_vu_5_8})



def danh_gia(request, pk, id_khach_hang):
    # Kiểm tra nếu employee_id có trong session
    if 'employee_id' not in request.session:
        # Nếu không có, chuyển hướng về trang nhân viên hoặc trang đăng nhập nếu cần
        return redirect('customer:nhan-vien', assigned_service_id=pk)
    
    # Lấy employee từ session
    employee_id = request.session['employee_id']
    print(employee_id)
    employee = Employee.objects.get(pk=employee_id)
    print(employee)
    # Tìm khách hàng dựa vào các tiêu chí lọc
    khach_hang = KhachHang.objects.filter(
        name__isnull=False,
        is_called=False,
        # employee=employee_id,  # Dùng employee_id từ session
        pk=id_khach_hang
    ).first()
    print(khach_hang)
    # Xác định URL để điều hướng sau khi đánh giá
    url = reverse('customer:nhan-vien', args=[pk, employee_id])
    
    # Lấy tất cả đánh giá
    danh_gia = CapNhatDuLieu.objects.all()

    # Xóa hàng chờ nếu tồn tại
    try:
        hang_can_xoa = DanhSachCho.objects.get(id=id_khach_hang)
        hang_can_xoa.delete()
    except ObjectDoesNotExist:
        hang_can_xoa = None

    # Xử lý POST request
    if khach_hang and request.method == 'POST':
        # Lấy đánh giá từ POST data
        feedback = request.POST.get('feedback')
        
        # Đặt is_called và is_calling thành True, lưu đánh giá cho khách hàng
        first_customer = KhachHang.objects.filter(id=id_khach_hang).first()
        if first_customer:
            first_customer.is_calling = True
            first_customer.is_called = True
            first_customer.employee = employee_id
            first_customer.feedback = feedback
            
            first_customer.save()

            # Tạo bản ghi trong ThongKe
            thong_ke = ThongKe.objects.create(
                ticket_number=first_customer.ticket_number,
                name=first_customer.name,
                dia_chi=first_customer.dia_chi,
                feedback=first_customer.feedback,
                service=first_customer.service,
                ngay_sinh=first_customer.ngay_sinh,
                so_cccd=first_customer.so_cccd,
                ngay_quet=timezone.now()  # Lưu thời gian hiện tại
            )

            # Xóa khách hàng khỏi danh sách chờ
            try:
                hang_can_xoa = DanhSachCho.objects.get(id=id_khach_hang)
                hang_can_xoa.delete()
            except ObjectDoesNotExist:
                pass

            # Điều hướng lại trang nhân viên đúng theo ID
            return redirect(url)

    # Điều hướng lại nếu không tìm thấy khách hàng
    elif not khach_hang:
        return redirect(url)

    # Render template nếu có khách hàng để đánh giá
    return render(request, 'customer/danh_gia_html', {
        'khach_hang': khach_hang,
        'danh_gia': danh_gia,
        'url': url,
    })

from django.db.models import Max
from django.db import transaction

def get_next_ticket_number(quay_so):
    """Delegate to customer.services."""
    return customer_services.get_next_ticket_number(quay_so)

def generate_qr_code(quay_so, ticket_number):
    """Delegate to customer.services."""
    return customer_services.generate_qr_code(quay_so, ticket_number)

@csrf_exempt
def get_ticket(request, quay_so):
    try:
        khachhang = ''
        print(f"Getting ticket for service quay_so={quay_so}")
        dich_vu = Service.objects.get(quay_so=quay_so)
        
        if request.method == 'POST':
            data = json.loads(request.body)
            if data.get('lay_so_truc_tiep', False):
                print("Processing direct ticket")
                next_ticket_number = get_next_ticket_number(quay_so)
                
                khachhang = KhachHang.objects.create(
                    name='',  
                    service=dich_vu,
                    ticket_number=next_ticket_number,
                    is_calling=False
                )
                
                print(f"Created customer with ticket {khachhang.ticket_number}")
                
                # Tạo QR code
                img_name = generate_qr_code(
                    quay_so=quay_so,
                    ticket_number=khachhang.ticket_number
                )
                
                try:
                    print_ticket(
                        ticket_number=str(khachhang.ticket_number),
                        counter=str(dich_vu.name),
                    )
                    return JsonResponse({
                        'success': True,
                        'so_thu_tu': khachhang.ticket_number,
                        'ho_va_ten': '',
                        'dia_chi': 'Không có',
                        'dich_vu': dich_vu.name,
                        'img_name': img_name,
                        'audio_url': None
                    })
                except:
                    canh_bao = 'MÁY IN GIẤY ĐANG BẢO TRÌ . VUI LÒNG NGỒI CHỜ GỌI SỐ'
                    return JsonResponse({
                        'success': True,
                        'so_thu_tu': khachhang.ticket_number,
                        'ho_va_ten': '',
                        'dia_chi': 'Không có',
                        'dich_vu': dich_vu.name,
                        'img_name': img_name,
                        'audio_url': None,
                        'canh_bao':canh_bao
                    })
  
            # Trường hợp nhập thông tin
            thong_tin = data.get('thong_tin', '').strip()
            if thong_tin:
                thong_tin = thong_tin.replace('|', "*").split('*')
                so_cccd = thong_tin[0]
                ho_va_ten = thong_tin[2].upper()
                ngay_sinh = datetime.strptime(thong_tin[3], "%d%m%Y").strftime("%d/%m/%Y")
                gioi_tinh = thong_tin[4]
                dia_chi = thong_tin[5]
                ngay_cap = datetime.strptime(thong_tin[6], "%d%m%Y").strftime("%d/%m/%Y")
                thanh_pho = dia_chi.replace('\x00', '').split(',')[-1].replace('\x00', '')
                ho_ten_cha = thong_tin[8] if len(thong_tin) > 8 else ''
                ho_ten_me = thong_tin[9] if len(thong_tin) > 9 else ''

                # Tạo khách hàng mới
                next_ticket_number = get_next_ticket_number(quay_so)
                khachhang = KhachHang.objects.create(
                    name=ho_va_ten,
                    dia_chi=dia_chi,
                    so_cccd=so_cccd,
                    ngay_sinh=ngay_sinh,
                    gioi_tinh=gioi_tinh,
                    ngay_cap=ngay_cap,
                    service=dich_vu,
                    ticket_number=next_ticket_number,
                    is_calling=False,
                    ho_ten_cha=ho_ten_cha,
                    ho_ten_me=ho_ten_me
                )

                # Sử dụng hàm generate_qr_code thay vì tạo trực tiếp
                img_name = generate_qr_code(
                    quay_so=quay_so,
                    ticket_number=khachhang.ticket_number
                )

                # Tạo file âm thanh phản hồi
                response_text = f"Cảm ơn {'Ông' if gioi_tinh == 'Nam' else 'Bà'} {ho_va_ten}, vui lòng theo dõi số thứ tự trên màn hình."
                tts = gTTS(response_text, lang='vi')
                audio_file = 'phan_hoi_lay_so.mp3'
                audio_path = os.path.join('media/audio', audio_file)
                tts.save(audio_path)

                return JsonResponse({
                    'success': True,
                    'so_thu_tu': khachhang.ticket_number,
                    'ho_va_ten': ho_va_ten,
                    'dia_chi': dia_chi,
                    'dich_vu': dich_vu.name,
                    'img_name': img_name,
                    'audio_url': f'/media/audio/{audio_file}'
                })

        return JsonResponse({'success': False, 'message': 'Yêu cầu không hợp lệ.'}, status=400)

    except Exception as e:
        print(f"Lỗi: {e}")
        return JsonResponse({'success': False, 'message': 'Có lỗi xảy ra, vui lòng thử lại.'}, status=500)

# ── Printer & Latin1 đã chuyển sang core/hardware/printer.py ──
# Sử dụng: from core.hardware.printer import print_ticket, convert_to_latin1


def chi_tiet_khach_hang(request, khachhang_id):
    khach_hang = get_object_or_404(KhachHang, pk=khachhang_id)
    return render(request, 'customer/chi_tiet_khach_hang.html', {'khach_hang': khach_hang})


import openpyxl
from django.http import HttpResponse
from .models import KhachHang

def export_khach_hang_to_excel(request, khachhang_id):
    khach_hang = get_object_or_404(KhachHang, pk=khachhang_id)

    # Tạo workbook và worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chi tiết khách hàng"

    # Thêm tiêu đề cột
    ws.append(['Tên nười dân', 'Địa chỉ', 'Số CCCD', 'Ngày sinh', 'Giới tính', 'Ngày cấp', 'Họ tên cha', 'Họ tên mẹ',])

    # Thêm dữ liệu khách hàng
    ws.append([
        khach_hang.name,
        khach_hang.dia_chi,
        khach_hang.so_cccd,
        khach_hang.ngay_sinh,
        khach_hang.gioi_tinh,
        khach_hang.ngay_cap,
        khach_hang.ho_ten_cha,
        khach_hang.ho_ten_me,

    ])

    # Tạo HTTP response với file Excel
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=khachhang_{khach_hang.id}.xlsx'
    wb.save(response)

    return response


def nhan_vien(request,assigned_service_id,id_nhan_vien):
    # Lấy thông tin nhân viên và dịch vụ được chỉ định
    employee = Employee.objects.get(pk=id_nhan_vien,assigned_service=assigned_service_id)
    request.session['employee_id'] = employee.pk

    assigned_service_name = employee.assigned_service.name

    # Trạng thái mặc định
    default_trangthai = TrangThai.objects.get(pk=1)

    # Lấy danh sách khách hàng và danh sách chờ
    khachhang = KhachHang.objects.filter(pk=assigned_service_id)
    tong_khach_hang = KhachHang.objects.filter(is_called=False)

    # Sắp xếp dịch vụ và số thứ tự khách hàng
    service_and_order = defaultdict(list)
    thu_tuc = DichVuTungQuay.objects.filter(service=assigned_service_id)
    danh_sach_cho = DanhSachCho.objects.filter(employee=assigned_service_id)

    for khach_hang in tong_khach_hang:
        if khach_hang.service:
            service_and_order[khach_hang.service].append(khach_hang.ticket_number)

    # Kiểm tra và cập nhật trạng thái mặc định cho khách hàng
    for khachhang in khachhang:
        if khachhang.trang_thai is None:
            KhachHang.objects.filter(trang_thai__isnull=True).update(trang_thai=default_trangthai)

    # Lấy danh sách khách hàng đang chờ
    waiting_customers = KhachHang.objects.filter(name__isnull=False, service=assigned_service_id, is_called=False,trang_thai=True)

    # Lấy thông tin dịch vụ
    service = Service.objects.get(pk=assigned_service_id)
    trang_thai = 'Không xác định'

    # Xác định trạng thái nhân viên và dịch vụ
    if employee.is_active:
        trang_thai = 'Hoạt động'
    else:
        trang_thai = 'Không hoạt động'

    # Xử lý yêu cầu POST để cập nhật trạng thái nhân viên và dịch vụ
    if request.method == 'POST' and 'scan_data_1' in request.POST:
        # Cập nhật trạng thái nhân viên và dịch vụ
        if employee.is_active:
            employee.is_active = False
            service.is_active = False
            trang_thai = 'Không hoạt động'
        else:
            employee.is_active = True
            service.is_active = True
            trang_thai = 'Hoạt động'

        # Lưu trạng thái mới
        employee.save()
        service.save()

        # Phát trạng thái qua WebSocket đến các client
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            'status_updates',  # Nhóm WebSocket mà các client lắng nghe
            {
                'type': 'update_status',
                'message': f'Service ID: {assigned_service_id}, Trạng thái: {trang_thai}'
            }
        )

        # Phản hồi lại client với trạng thái mới
        

    # Đếm số khách hàng đang chờ
    count = waiting_customers.count()

    # Render giao diện với dữ liệu đã chuẩn bị
    return render(request, 'customer/nhan_vien_hanh_chinh.html', {
        'loai_dich_vu': assigned_service_name,
        'ten_nhan_vien': employee.name,
        'khach_hang_cho': waiting_customers,
        'employee': employee,
        'dem': count,
        'trang_thai': trang_thai,
        'quay': employee.pk,
        'danh_sach_cho': danh_sach_cho,
        'tong_khach_hang': tong_khach_hang,
        'service': service_and_order.items(),
        'thu_tuc': thu_tuc
    })

def danh_sach_nhan_vien(request,assigned_service_id):
    nhan_vien= Employee.objects.filter(assigned_service_id=assigned_service_id)
    dich_vu_can_bo = Service.objects.get(pk=assigned_service_id)
    return render(request, 'customer/danh_sach_nhan_vien.html',
    {'danh_sach_nhan_vien':nhan_vien,
     'danh_sach_quay':dich_vu_can_bo,
    })



def bang_khach_hang(request):
    dich_vu= Service.objects.all
    khach_hang_list = KhachHang.objects.filter(name__isnull=False,is_called=False,trang_thai=True)
    grouped_khach_hang = defaultdict(list)
    for khach_hang in khach_hang_list:
        # Group khach_hang theo dịch vụ
        grouped_khach_hang[khach_hang.service].append(khach_hang)
    
    # Chuyển từ điển thành danh sách tuples để truyền vào template
    grouped_khach_hang_list = list(grouped_khach_hang.items())
    
    
    return render(request,'customer/bang_khach_hang.html',{
                  'grouped_khach_hang_list': grouped_khach_hang_list,
                      
                         'dich_vu':dich_vu}
                        )

     
    
def gui_khach_hang(request,pk,ticket_number):
    
    so_thu_tu = KhachHang.objects.filter(name__isnull=False, is_called=False, feedback=None,service_id=pk).order_by('ticket_number')
    ho_va_ten=KhachHang.objects.filter(ticket_number=ticket_number,employee=pk,feedback=None).first()
    first_ticket_number = so_thu_tu.first().ticket_number
    danh_sach_so_thu_tu = KhachHang.objects.filter(is_called=False, employee=pk ).values_list('ticket_number', flat=True)
    danh_sach_so_thu_tu = [int(stt) for stt in danh_sach_so_thu_tu]
    # Đếm số lượng số thứ tự liền trước
    so_luong_lien_truoc = len([stt for stt in danh_sach_so_thu_tu if stt < ticket_number])


    if ticket_number == first_ticket_number:
        dem = 0
    else:
        if ho_va_ten:  
            dem = so_luong_lien_truoc
            if dem < 0:
                dem = 'Đã xử lý xong'
        else:
            dem=''
    khach_hang = KhachHang.objects.filter(name__isnull=False,is_called=False, employee=pk, ticket_number=ticket_number)
    thoi_gian = datetime.now()
    time = thoi_gian.strftime("%H:%M:%S")
    
    
    return render(request,'customer/gui_khach_hang.html',{
                            'khach_hang': khach_hang,
                            'thoi_gian' :time,
                            'dem': dem,

                         
                         } )
    
        
# def generate_qr_code(request,pk,ticket_number):
#     # Tạo đường dẫn URL với các tham số
#     context = {}
#     if request.method == "POST":
#         data = request.POST['qr_text']
#         img = make(data)
#         test='a'
#         img_name = f'{test}.png'
#         img.save(settings.MEDIA_ROOT + '/' + img_name)
#         context = {
#             'img_name': img_name
#         }
        
       

#     return render(request, "Qrcodes/customer/qr.html",context)

from datetime import datetime
from django.shortcuts import redirect
from openpyxl import Workbook
from django.db import connection
from .models import KhachHang, Service

from datetime import datetime
from django.shortcuts import redirect
from openpyxl import Workbook
from django.db import connection
from .models import KhachHang, Service

def tong_hop(request):
    if 's_khachhang' not in request.session:
        return redirect('customer:dang-nhap')
    
    selected_columns = ['name', 'ticket_number', 'service_id', 'feedback']
    
    # Fetch data from the customer table
    with connection.cursor() as cursor:
        cursor.execute(f"SELECT {', '.join(selected_columns)} FROM customer_khachhang")
        columns = [col[0] for col in cursor.description]
        data = cursor.fetchall()

    # Create Workbook and Worksheet
    wb = Workbook()
    ws = wb.active

    # Add column names to the Worksheet
    ws.append(selected_columns)

    # Add data rows to the Worksheet
    for row in data:
        ws.append(row)

    # Save the Workbook to an Excel file
    current_date = datetime.now().strftime('%Y-%m-%d')
    file_name = f'output_{current_date}.xlsx'
    folder_report = MEDIA_ROOT + 'QRcodes\\reports\\'
    path_report = folder_report + file_name
    wb.save(path_report)
    
    # Reset session and delete all customers
    if 's_khachhang' not in request.session:
        return redirect('customer:dang-nhap')
    
    KhachHang.objects.all().delete()

    # Update all records where 'da_phuc_vu' is True
    KhachHang.objects.update(is_called=False)

    # Retrieve all active services (queues) with prefixes
    services = Service.objects.filter(is_active=True)
   # Tạo tối đa 99 vé cho mỗi quầy
    

    # for service in services:
    #     prefix = service.prefix  # e.g., 1000 for prefix 1, 2000 for prefix 2
    #     print(prefix)
       
    #     KhachHang.objects.create(ticket_number=prefix, is_called=False, is_calling=False, service=service)
    
    # Clear session variable
    # del request.session['s_khachhang']
    
    return redirect('customer:trang-chu')





def text_to_speech(request):
    """
    HTTP endpoint: đọc văn bản qua TTS.
    Delegates to core.hardware.audio.speak()
    """
    if request.method == 'POST':
        text_to_read = request.POST.get('text', '')
        result = speak(text_to_read)
        return JsonResponse(result)
    return HttpResponse('', content_type='audio/mpeg')


def dem(request):
    danh_sach = ThongKe.objects.all()
    du_lieu_feedback = CapNhatDuLieu.objects.all()

    # Lặp qua từng đối tượng trong du_lieu_feedback và đếm số lượng phản hồi cho mỗi id
    for du_lieu in du_lieu_feedback:
        id = du_lieu.id
        count = 0
        for feedback in danh_sach:
            if feedback.feedback == du_lieu.feedback:
                count += 1

        # Kiểm tra xem đã tồn tại bản ghi cho id trong FeedbackCount hay chưa
        try:
            feedback_count = CapNhatDuLieu.objects.get(id=id)
            feedback_count.tong = count
            feedback_count.save()  # Cập nhật bản ghi đã tồn tại
        except CapNhatDuLieu.DoesNotExist:
            CapNhatDuLieu.objects.create(id=id, count=count)  # Tạo mới bản ghi nếu chưa tồn tại
    # Gửi số lượng phản hồi cho mỗi id đến giao diện
 





def feedback_chart_data(request):
    # Lấy dữ liệu feedback theo ngày tháng, sắp xếp theo ngày tháng
    feedback_data = ThongKe.objects.filter(ngay_quet__isnull=False).values('ngay_quet__date', 'feedback').annotate(count=Count('id')).order_by('ngay_quet__date')

    # Tạo danh sách để lưu thông tin feedback cho mỗi ngày
    data = []
    for item in feedback_data:
        date_obj = item['ngay_quet__date']
        # Định dạng lại ngày tháng
        formatted_date = date_obj.strftime('%d-%b-%Y')

        # Tạo một đối tượng mới để lưu thông tin feedback
        feedback_item = {
            'date': formatted_date,
            'count': item['count'],
            'feedbacks': {item['feedback']: item['count']}
        }

        # Kiểm tra xem ngày đã tồn tại trong danh sách hay chưa
        existing_item = next((x for x in data if x['date'] == formatted_date), None)
        if existing_item:
            existing_item['count'] += item['count']
            if item['feedback'] not in existing_item['feedbacks']:
                existing_item['feedbacks'][item['feedback']] = item['count']
            else:
                existing_item['feedbacks'][item['feedback']] += item['count']
        else:
            data.append(feedback_item)

    return JsonResponse({'data': data})

def feedback(request):
    return render(request, 'customer/feedback_chart.html')


def move_to_waiting_list(request, khachhang_id):
    khachhang = get_object_or_404(KhachHang, ticket_number=khachhang_id)
    print(khachhang.trang_thai)
    employee_id = request.session['employee_id']
    
    if request.method == "POST":
        khachhang = get_object_or_404(KhachHang, ticket_number=khachhang_id)
        print(khachhang)
        nhan_vien=khachhang.employee
        url = reverse('customer:nhan-vien', args=[khachhang.service.id,employee_id])
        if khachhang:
            khachhang.trang_thai = False
            khachhang.save()
            DanhSachCho.objects.create(khach_hang=khachhang,name=khachhang.name,ticket_number=khachhang_id,employee=khachhang.employee)
            redirect(url)
        else:
            return JsonResponse({'error': 'Khach hang khong dang duoc goi.'}, status=400)
    
    return  redirect(url)
    
    
    
def lay_so(request):
    context = {}
    img_name={}
    data_json_1=''      
    ho_va_ten=''
    so_thu_tu='' 
    dia_chi=''
    sothutu=KhachHang()
    thoi_gian_1= Manage.objects.get(id=1)
    thoi_gian_1=thoi_gian_1.time
    
    if request.method == 'POST' and 'scan_data_1' in request.POST:
        sothutu = KhachHang.objects.all()
        thoi_gian_1= Manage.objects.get(id=1)
        thoi_gian_1=thoi_gian_1.time
        sothutu = KhachHang.objects.filter(is_calling=False).first()
        thong_tin=request.POST.get('thong_tin')
        try :
            thong_tin = thong_tin.rstrip()
            thong_tin=thong_tin.replace('|',"*")
            thong_tin=thong_tin.split('*')
            so_cccd = thong_tin[0]
            ho_va_ten = thong_tin[2].upper()
            ngay_sinh = thong_tin[3]
            ngay_sinh= datetime.strptime(ngay_sinh, "%d%m%Y")
            ngay_sinh = datetime.strftime(ngay_sinh, "%d/%m/%Y")
            gioi_tinh = thong_tin[4]
            dia_chi = thong_tin[5]
            ngay_cap = thong_tin[6]
            ngay_cap= datetime.strptime(ngay_cap, "%d%m%Y")
            ngay_cap = datetime.strftime(ngay_cap, "%d/%m/%Y")
            thanh_pho = dia_chi.replace('\x00','').split(',')
            thanh_pho_chuan=thanh_pho[-1]
            thoi_gian = datetime.now()
            thoi_gian= thoi_gian.strftime("%d/%m/%Y %H:%M")
            
            so_thu_tu = sothutu.ticket_number
            data_json_1 = {
            'ho_va_ten':ho_va_ten,
            'so_cccd': so_cccd,
            'dia_chi':dia_chi,
            'so_thu_tu': so_thu_tu,
            'ngay_sinh': ngay_sinh,
            'gioi_tinh':gioi_tinh.replace('\x00',''),
            'dia_chi': dia_chi.replace('\x00',''),
            'ngay_cap':ngay_cap.replace('\x00',''),
            'thanh_pho': thanh_pho_chuan.replace('\x00',''),
            'so_thu_tu':so_thu_tu
            
        }
        except:
            thong_bao = '''
                <div class="alert alert-success text-center display-3" role="alert">
                    Vui lòng chọn đúng mã QR trên căn cước công dân hoặc VneID
                </div>
                '''
            return render(request, 'customer/lay_so_1.html',{'thong_bao':thong_bao,'thoi_gian':thoi_gian_1})      
        
        request.session['scan_data_1'] = data_json_1
        context = {}
         
    return render(request, 'customer/lay_so_1.html',{'so_thu_tu': sothutu.ticket_number,
                                                           'data_1': data_json_1,
                                                           'ho_va_ten':ho_va_ten,
                                                           'dia_chi':dia_chi,
                                                            'so_thu_tu': so_thu_tu,
                                                            'context':context,
                                                            'img_name':img_name,
                                                            'thoi_gian':thoi_gian_1}
                  )



def feedback_statistics(request):
    # Lấy dữ liệu feedback và thống kê số lượng phản hồi
    feedback_data = ThongKe.objects.filter(ngay_quet__isnull=False).values('feedback').annotate(count=Count('id'))

    # Tạo danh sách để lưu thông tin phản hồi và số lượng của chúng
    statistics = []
    for item in feedback_data:
        statistics.append({
            'feedback': item['feedback'],
            'count': item['count']
        })

    return JsonResponse({'statistics': statistics})



def get_all_status(request):
    employees = Employee.objects.all()
    services = Service.objects.all()
    
    response_data = []
    for employee in employees:
        service = services.get(pk=employee.pk)
        response_data.append({
            'employee': employee.pk,
            'employee_status': employee.is_active,
            'service_status': service.is_active,
        })
    
    return JsonResponse(response_data, safe=False)




# views.py
from django.shortcuts import render, redirect
from .models import Banner
from .forms import BannerForm


def manage_banners(request, banner_id=None):
    if banner_id:
        banner = get_object_or_404(Banner, id=banner_id)
    else:
        banner = None

    if request.method == 'POST':
        if 'delete' in request.POST:
            if banner:
                banner.delete()
                return redirect('customer:banner_list')
        else:
            form = BannerForm(request.POST, request.FILES, instance=banner)
            if form.is_valid():
                form.save()
                return redirect('customer:banner_list')
    else:
        form = BannerForm(instance=banner)
    
    banners = Banner.objects.all()
    return render(request, 'customer/manage_banners.html', {'form': form, 'banners': banners, 'banner': banner})


def show_thong_ke(request):
    return render(request, 'customer/change_list_graph.html')


from .models import KhachHang, Service




def create_khachhang(request):
    services = Service.objects.all()  # Lấy tất cả các dịch vụ
    if request.method == 'POST':
        name = request.POST.get('name')
        dia_chi = request.POST.get('dia_chi')
        so_cccd = request.POST.get('so_cccd')
        ngay_sinh = request.POST.get('ngay_sinh')
        gioi_tinh = request.POST.get('gioi_tinh')
        ngay_cap = request.POST.get('ngay_cap')
        service_id = request.POST.get('service')  # Dịch vụ khách hàng chọn

        # Lấy số thứ tự tiếp theo dựa trên dịch vụ
        next_ticket_number = get_next_ticket_number(service_id)

        # Lấy đối tượng Service
        service = Service.objects.get(id=service_id)

        # Tạo khách hàng mới
        khachhang = KhachHang.objects.create(
            name=name,
            dia_chi=dia_chi,
            so_cccd=so_cccd,
            ngay_sinh=ngay_sinh,
            gioi_tinh=gioi_tinh,
            ngay_cap=ngay_cap,
            service=service,  # Gán dịch vụ
            ticket_number=next_ticket_number
        )
        khachhang.save()

        return render(request, 'customer/khachhang_success.html', {'khachhang': khachhang})

    return render(request, 'customer/create_khachhang.html', {'services': services})


def service_employee_management_ajax(request):
    # Check if the request is an AJAX request using the header
    if request.headers.get('x-requested-with') == 'XMLHttpRequest':
        
        # Handle delete service request
        if 'delete_service' in request.POST:
            service_id = request.POST.get('service_id')
            service = get_object_or_404(Service, id=service_id)
            service.delete()
            return JsonResponse({'success': True, 'service_id': service_id})

        # Handle delete employee request
        if 'delete_employee' in request.POST:
            employee = request.POST.get('employee')
            employee = get_object_or_404(Employee, id=employee)
            employee.delete()
            return JsonResponse({'success': True, 'employee': employee})

        # Handle service update request
        if 'service_form' in request.POST:
            service_id = request.POST.get('service_id')
            service_name = request.POST.get('name')
            service = get_object_or_404(Service, id=service_id)

            # Update service details
            service.name = service_name
            service.save()
            return JsonResponse({'success': True, 'service': {'id': service.id, 'name': service.name}})

        # Handle employee update request
        if 'employee_form' in request.POST:
            employee = request.POST.get('employee')
            employee_name = request.POST.get('name')
            employee_position = request.POST.get('position')
            employee = get_object_or_404(Employee, id=employee)

            # Update employee details
            employee.name = employee_name
            employee.position = employee_position
            employee.save()
            return JsonResponse({'success': True, 'employee': {'id': employee.id, 'name': employee.name, 'position': employee.position}})
    
    # If it's not an AJAX request, render the HTML template
    else:
        services = Service.objects.all()  # Fetch all services from the database
        employees = Employee.objects.all()  # Fetch all employees from the database
        return render(request, 'customer/service_employee_management.html', {
            'services': services,
            'employees': employees
        })


def lam_thu_tuc(request):
    """Trang làm thủ tục toàn trình."""
    return render(request, 'customer/lam_thu_tuc.html')


def dang_ky_ket_hon(request):
    """Workspace làm thủ tục đăng ký kết hôn toàn trình."""
    return render(request, 'customer/dang_ky_ket_hon.html')
